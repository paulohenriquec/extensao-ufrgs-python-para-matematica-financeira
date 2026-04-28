#%%
from openpyxl import load_workbook
import statistics as st
import numpy as np
import matplotlib.pyplot as plt

wb = load_workbook('dados.xlsx', data_only=True)
p = wb['Planilha1']

linha = p.max_row
coluna = p.max_column

dados = []

for i in range(1, linha+1):
    dados.append(p.cell(row=i, column=2).value)

for i in range(len(dados)):
    dados[i] = float(dados[i])

media = st.mean(dados)
desvio = st.pstdev(dados)

plt.subplot(211)
plt.hist(dados, bins=7)
plt.xlabel('disciplina')
plt.ylabel('nota')

plt.subplot(212)
plt.boxplot(dados)